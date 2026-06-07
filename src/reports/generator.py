import logging
import os
from datetime import date, datetime, timedelta
from typing import Dict, List, Optional
from collections import defaultdict
from sqlalchemy import and_, func

from src.database import (
    get_db, Transaction, AccountBalance, ExchangeExecution,
    FundGap, RiskExposure, MonthlyReport
)
from src.banking.exchange_rates import ExchangeRateManager
from src.utils.helpers import generate_id, round_currency, safe_json_dumps

logger = logging.getLogger(__name__)

class ReportGenerator:
    def __init__(self, rate_manager: ExchangeRateManager, 
                 currencies: List[str], base_currency: str = 'CNY'):
        self.rate_manager = rate_manager
        self.currencies = currencies
        self.base_currency = base_currency
        self.ensure_export_dirs()
    
    def ensure_export_dirs(self):
        os.makedirs('exports/pdf', exist_ok=True)
        os.makedirs('exports/excel', exist_ok=True)
    
    def generate_monthly_report(self, year: int, month: int) -> Dict:
        logger.info(f"Generating monthly report for {year}-{month:02d}")
        
        report_month = f"{year}-{month:02d}"
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)
        
        report_data = {
            'report_month': report_month,
            'summary': self._calculate_summary(start_date, end_date),
            'fund_collection': self._calculate_fund_collection(start_date, end_date),
            'exchange_costs': self._calculate_exchange_costs(start_date, end_date),
            'fx_impact': self._calculate_fx_impact(start_date, end_date),
            'risk_metrics': self._calculate_risk_metrics(end_date),
            'generated_at': datetime.now().isoformat()
        }
        
        report_id = generate_id('RPT')
        
        excel_path = self._export_to_excel(report_id, report_data)
        pdf_path = self._export_to_pdf(report_id, report_data)
        
        self._save_report(report_id, report_month, report_data, excel_path, pdf_path)
        
        logger.info(f"Monthly report generated: {report_id}")
        return {
            'report_id': report_id,
            'report_month': report_month,
            'excel_path': excel_path,
            'pdf_path': pdf_path,
            'report_data': report_data
        }
    
    def _calculate_summary(self, start_date: date, end_date: date) -> Dict:
        with get_db() as db:
            total_inflow = 0
            total_outflow = 0
            txn_count = 0
            
            for currency in self.currencies:
                inflow = db.query(func.sum(Transaction.amount)).filter(
                    and_(
                        Transaction.currency == currency,
                        Transaction.amount > 0,
                        Transaction.txn_date >= start_date,
                        Transaction.txn_date <= end_date
                    )
                ).scalar() or 0
                
                outflow = db.query(func.sum(Transaction.amount)).filter(
                    and_(
                        Transaction.currency == currency,
                        Transaction.amount < 0,
                        Transaction.txn_date >= start_date,
                        Transaction.txn_date <= end_date
                    )
                ).scalar() or 0
                
                total_inflow += self.rate_manager.convert(inflow, currency, self.base_currency, end_date)
                total_outflow += self.rate_manager.convert(abs(outflow), currency, self.base_currency, end_date)
                
                txn_count += db.query(func.count(Transaction.id)).filter(
                    and_(
                        Transaction.currency == currency,
                        Transaction.txn_date >= start_date,
                        Transaction.txn_date <= end_date
                    )
                ).scalar() or 0
            
            return {
                'total_inflow_base': round_currency(total_inflow, self.base_currency),
                'total_outflow_base': round_currency(total_outflow, self.base_currency),
                'net_flow_base': round_currency(total_inflow - total_outflow, self.base_currency),
                'transaction_count': txn_count,
                'currencies_used': len(self.currencies)
            }
    
    def _calculate_fund_collection(self, start_date: date, end_date: date) -> Dict:
        with get_db() as db:
            collection_rates = {}
            
            for currency in self.currencies:
                expected_inflow = db.query(func.sum(Transaction.amount)).filter(
                    and_(
                        Transaction.currency == currency,
                        Transaction.amount > 0,
                        Transaction.txn_date >= start_date,
                        Transaction.txn_date <= end_date
                    )
                ).scalar() or 0
                
                actual_balance = 0
                latest_balance = db.query(AccountBalance).filter(
                    and_(
                        AccountBalance.currency == currency,
                        AccountBalance.balance_date <= end_date
                    )
                ).order_by(AccountBalance.balance_date.desc()).first()
                
                if latest_balance:
                    actual_balance = latest_balance.balance
                
                if expected_inflow > 0:
                    collection_rate = min(1.0, actual_balance / expected_inflow)
                else:
                    collection_rate = 1.0
                
                collection_rates[currency] = {
                    'expected_inflow': round_currency(expected_inflow, currency),
                    'actual_balance': round_currency(actual_balance, currency),
                    'collection_rate': round(collection_rate * 100, 2)
                }
            
            return collection_rates
    
    def _calculate_exchange_costs(self, start_date: date, end_date: date) -> Dict:
        with get_db() as db:
            executions = db.query(ExchangeExecution).filter(
                and_(
                    ExchangeExecution.executed_at >= start_date,
                    ExchangeExecution.executed_at <= end_date,
                    ExchangeExecution.status == 'completed'
                )
            ).all()
            
            total_fee = 0
            total_volume = 0
            
            by_pair = defaultdict(lambda: {'count': 0, 'volume': 0, 'fee': 0})
            
            for exec in executions:
                fee_base = self.rate_manager.convert(
                    exec.fee, exec.target_amount and 'CNY' or exec.source_account_id[:3], 
                    self.base_currency, exec.executed_at.date() if exec.executed_at else end_date
                )
                volume_base = self.rate_manager.convert(
                    exec.source_amount, exec.proposal_id and 'USD' or exec.source_account_id[:3],
                    self.base_currency, exec.executed_at.date() if exec.executed_at else end_date
                )
                
                total_fee += fee_base
                total_volume += volume_base
                
                pair = f"{exec.source_account_id[:3] if exec.source_account_id else 'XXX'}/{exec.target_account_id[:3] if exec.target_account_id else 'XXX'}"
                by_pair[pair]['count'] += 1
                by_pair[pair]['volume'] += volume_base
                by_pair[pair]['fee'] += fee_base
            
            avg_cost_rate = (total_fee / total_volume * 100) if total_volume > 0 else 0
            
            return {
                'total_exchanges': len(executions),
                'total_volume_base': round_currency(total_volume, self.base_currency),
                'total_fees_base': round_currency(total_fee, self.base_currency),
                'avg_cost_rate_percent': round(avg_cost_rate, 4),
                'by_pair': {k: {
                    'count': v['count'],
                    'volume_base': round_currency(v['volume'], self.base_currency),
                    'fee_base': round_currency(v['fee'], self.base_currency)
                } for k, v in by_pair.items()}
            }
    
    def _calculate_fx_impact(self, start_date: date, end_date: date) -> Dict:
        impact_by_currency = {}
        
        for currency in self.currencies:
            if currency == self.base_currency:
                continue
            
            start_rate = self.rate_manager.get_rate(currency, self.base_currency, start_date)
            end_rate = self.rate_manager.get_rate(currency, self.base_currency, end_date)
            
            rate_change = (end_rate - start_rate) / start_rate if start_rate != 0 else 0
            
            avg_balance = self._get_average_balance(currency, start_date, end_date)
            avg_balance_base = self.rate_manager.convert(avg_balance, currency, self.base_currency, end_date)
            
            impact = avg_balance_base * rate_change
            
            impact_by_currency[currency] = {
                'start_rate': start_rate,
                'end_rate': end_rate,
                'rate_change_percent': round(rate_change * 100, 2),
                'avg_balance': round_currency(avg_balance, currency),
                'fx_impact_base': round_currency(impact, self.base_currency)
            }
        
        return impact_by_currency
    
    def _get_average_balance(self, currency: str, start_date: date, end_date: date) -> float:
        with get_db() as db:
            balances = db.query(AccountBalance).filter(
                and_(
                    AccountBalance.currency == currency,
                    AccountBalance.balance_date >= start_date,
                    AccountBalance.balance_date <= end_date
                )
            ).all()
            
            if not balances:
                latest = db.query(AccountBalance).filter(
                    AccountBalance.currency == currency
                ).order_by(AccountBalance.balance_date.desc()).first()
                return latest.balance if latest else 0
            
            return sum(b.balance for b in balances) / len(balances)
    
    def _calculate_risk_metrics(self, end_date: date) -> Dict:
        with get_db() as db:
            exposures = db.query(RiskExposure).filter(
                RiskExposure.exposure_date == end_date
            ).all()
            
            if not exposures:
                return {'total_exposure_base': 0, 'by_currency': {}}
            
            total_exposure = sum(e.exposure_amount for e in exposures)
            
            return {
                'total_exposure_base': round_currency(total_exposure, self.base_currency),
                'by_currency': {
                    e.currency: {
                        'net_position': e.net_position,
                        'exposure_base': round_currency(e.exposure_amount, self.base_currency),
                        'var_95_base': round_currency(e.var_95, self.base_currency) if e.var_95 else 0
                    }
                    for e in exposures
                }
            }
    
    def _export_to_excel(self, report_id: str, report_data: Dict) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            wb = openpyxl.Workbook()
            
            ws1 = wb.active
            ws1.title = "摘要"
            
            headers = ['指标', '数值']
            for col, header in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            
            summary = report_data['summary']
            row = 2
            for key, value in summary.items():
                ws1.cell(row=row, column=1, value=key)
                ws1.cell(row=row, column=2, value=value)
                row += 1
            
            ws2 = wb.create_sheet("资金归集")
            ws2.cell(row=1, column=1, value="币种").font = Font(bold=True)
            ws2.cell(row=1, column=2, value="预期流入").font = Font(bold=True)
            ws2.cell(row=1, column=3, value="实际余额").font = Font(bold=True)
            ws2.cell(row=1, column=4, value="归集率(%)").font = Font(bold=True)
            
            row = 2
            for curr, data in report_data['fund_collection'].items():
                ws2.cell(row=row, column=1, value=curr)
                ws2.cell(row=row, column=2, value=data['expected_inflow'])
                ws2.cell(row=row, column=3, value=data['actual_balance'])
                ws2.cell(row=row, column=4, value=data['collection_rate'])
                row += 1
            
            excel_path = f"exports/excel/monthly_report_{report_id}.xlsx"
            wb.save(excel_path)
            return excel_path
            
        except ImportError:
            logger.warning("openpyxl not available, skipping Excel export")
            return ""
    
    def _export_to_pdf(self, report_id: str, report_data: Dict) -> str:
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            
            pdf_path = f"exports/pdf/monthly_report_{report_id}.pdf"
            doc = SimpleDocTemplate(pdf_path, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []
            
            title = f"资金池月度分析报告 - {report_data['report_month']}"
            elements.append(Paragraph(title, styles['Title']))
            elements.append(Spacer(1, 20))
            
            elements.append(Paragraph("一、总体概览", styles['Heading2']))
            summary = report_data['summary']
            summary_data = [
                ['总流入', f"{summary['total_inflow_base']:,.2f} {self.base_currency}"],
                ['总流出', f"{summary['total_outflow_base']:,.2f} {self.base_currency}"],
                ['净流入', f"{summary['net_flow_base']:,.2f} {self.base_currency}"],
                ['交易笔数', str(summary['transaction_count'])],
            ]
            t = Table(summary_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            elements.append(Spacer(1, 20))
            
            elements.append(Paragraph("二、换汇成本分析", styles['Heading2']))
            exchange_data = [
                ['换汇笔数', str(report_data['exchange_costs']['total_exchanges'])],
                ['总交易量', f"{report_data['exchange_costs']['total_volume_base']:,.2f}"],
                ['总费用', f"{report_data['exchange_costs']['total_fees_base']:,.2f}"],
                ['平均成本率', f"{report_data['exchange_costs']['avg_cost_rate_percent']}%"],
            ]
            t2 = Table(exchange_data)
            t2.setStyle(TableStyle([('GRID', (0, 0), (-1, -1), 1, colors.black)]))
            elements.append(t2)
            
            doc.build(elements)
            return pdf_path
            
        except ImportError:
            logger.warning("reportlab not available, skipping PDF export")
            return ""
    
    def _save_report(self, report_id: str, report_month: str, report_data: Dict,
                     excel_path: str, pdf_path: str):
        with get_db() as db:
            report = MonthlyReport(
                report_id=report_id,
                report_month=report_month,
                report_data=safe_json_dumps(report_data),
                pdf_path=pdf_path,
                excel_path=excel_path,
                generated_by='system'
            )
            db.add(report)
            db.commit()
