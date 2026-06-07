import logging
import os
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple
from sqlalchemy import and_, or_, func
import json

from src.database import get_db, Transaction, AccountBalance, ExchangeExecution
from src.utils.helpers import chunk_list

logger = logging.getLogger(__name__)

class TransactionQuery:
    def __init__(self, batch_size: int = 1000):
        self.batch_size = batch_size
    
    def query_transactions(self, 
                           start_date: date = None,
                           end_date: date = None,
                           currencies: List[str] = None,
                           account_ids: List[str] = None,
                           txn_type: str = None,
                           min_amount: float = None,
                           max_amount: float = None,
                           category: str = None,
                           counterparty: str = None,
                           page: int = 1,
                           page_size: int = 100) -> Dict:
        logger.info(f"Querying transactions with filters: {locals()}")
        
        with get_db() as db:
            query = db.query(Transaction)
            
            if start_date:
                query = query.filter(Transaction.txn_date >= start_date)
            if end_date:
                query = query.filter(Transaction.txn_date <= end_date)
            if currencies:
                query = query.filter(Transaction.currency.in_(currencies))
            if account_ids:
                query = query.filter(Transaction.account_id.in_(account_ids))
            if txn_type:
                query = query.filter(Transaction.txn_type == txn_type)
            if min_amount is not None:
                query = query.filter(func.abs(Transaction.amount) >= min_amount)
            if max_amount is not None:
                query = query.filter(func.abs(Transaction.amount) <= max_amount)
            if category:
                query = query.filter(Transaction.category.like(f'%{category}%'))
            if counterparty:
                query = query.filter(Transaction.counterparty.like(f'%{counterparty}%'))
            
            total_count = query.count()
            
            transactions = query.order_by(Transaction.txn_date.desc(), Transaction.id.desc()) \
                .offset((page - 1) * page_size) \
                .limit(page_size) \
                .all()
            
            result = [
                {
                    'txn_id': txn.txn_id,
                    'account_id': txn.account_id,
                    'currency': txn.currency,
                    'amount': txn.amount,
                    'txn_type': txn.txn_type,
                    'counterparty': txn.counterparty,
                    'description': txn.description,
                    'txn_date': txn.txn_date,
                    'txn_time': txn.txn_time,
                    'category': txn.category,
                    'status': txn.status,
                    'is_manual': txn.is_manual
                }
                for txn in transactions
            ]
            
            return {
                'total_count': total_count,
                'page': page,
                'page_size': page_size,
                'total_pages': (total_count + page_size - 1) // page_size,
                'data': result
            }
    
    def query_full_lifecycle(self, 
                             start_date: date,
                             end_date: date,
                             currencies: List[str] = None,
                             account_ids: List[str] = None) -> List[Dict]:
        logger.info(f"Querying full lifecycle transactions from {start_date} to {end_date}")
        
        all_transactions = []
        page = 1
        page_size = 5000
        
        while True:
            result = self.query_transactions(
                start_date=start_date,
                end_date=end_date,
                currencies=currencies,
                account_ids=account_ids,
                page=page,
                page_size=page_size
            )
            
            all_transactions.extend(result['data'])
            
            if page >= result['total_pages']:
                break
            
            page += 1
        
        logger.info(f"Retrieved {len(all_transactions)} transactions in full lifecycle query")
        return all_transactions
    
    def batch_export_to_excel(self, transactions: List[Dict], 
                              export_path: str) -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "交易流水"
            
            headers = ['交易ID', '账户', '币种', '金额', '类型', '交易对手', 
                       '描述', '交易日期', '交易时间', '分类', '状态']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
            
            row = 2
            for txn in transactions:
                ws.cell(row=row, column=1, value=txn['txn_id'])
                ws.cell(row=row, column=2, value=txn['account_id'])
                ws.cell(row=row, column=3, value=txn['currency'])
                ws.cell(row=row, column=4, value=txn['amount'])
                ws.cell(row=row, column=5, value=txn['txn_type'])
                ws.cell(row=row, column=6, value=txn['counterparty'])
                ws.cell(row=row, column=7, value=txn['description'])
                ws.cell(row=row, column=8, value=str(txn['txn_date']))
                ws.cell(row=row, column=9, value=str(txn['txn_time']) if txn['txn_time'] else '')
                ws.cell(row=row, column=10, value=txn['category'])
                ws.cell(row=row, column=11, value=txn['status'])
                row += 1
            
            os.makedirs(os.path.dirname(export_path), exist_ok=True)
            wb.save(export_path)
            logger.info(f"Exported {len(transactions)} transactions to {export_path}")
            return export_path
            
        except ImportError:
            logger.error("openpyxl not available for Excel export")
            raise
    
    def batch_export_to_csv(self, transactions: List[Dict], 
                            export_path: str) -> str:
        import csv
        
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        with open(export_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(['交易ID', '账户', '币种', '金额', '类型', '交易对手',
                           '描述', '交易日期', '交易时间', '分类', '状态'])
            
            for txn in transactions:
                writer.writerow([
                    txn['txn_id'],
                    txn['account_id'],
                    txn['currency'],
                    txn['amount'],
                    txn['txn_type'],
                    txn['counterparty'],
                    txn['description'],
                    str(txn['txn_date']),
                    str(txn['txn_time']) if txn['txn_time'] else '',
                    txn['category'],
                    txn['status']
                ])
        
        logger.info(f"Exported {len(transactions)} transactions to {export_path}")
        return export_path
    
    def batch_export_to_json(self, transactions: List[Dict],
                             export_path: str) -> str:
        os.makedirs(os.path.dirname(export_path), exist_ok=True)
        
        with open(export_path, 'w', encoding='utf-8') as f:
            json.dump(transactions, f, ensure_ascii=False, indent=2, default=str)
        
        logger.info(f"Exported {len(transactions)} transactions to {export_path}")
        return export_path
    
    def get_account_summary(self, start_date: date, end_date: date,
                            account_ids: List[str] = None) -> List[Dict]:
        with get_db() as db:
            query = db.query(
                Transaction.account_id,
                Transaction.currency,
                Transaction.txn_type,
                func.count(Transaction.id).label('txn_count'),
                func.sum(Transaction.amount).label('total_amount')
            ).filter(
                and_(
                    Transaction.txn_date >= start_date,
                    Transaction.txn_date <= end_date
                )
            )
            
            if account_ids:
                query = query.filter(Transaction.account_id.in_(account_ids))
            
            results = query.group_by(
                Transaction.account_id,
                Transaction.currency,
                Transaction.txn_type
            ).all()
            
            summary = {}
            for row in results:
                key = (row.account_id, row.currency)
                if key not in summary:
                    summary[key] = {
                        'account_id': row.account_id,
                        'currency': row.currency,
                        'inflow_count': 0,
                        'outflow_count': 0,
                        'inflow_amount': 0,
                        'outflow_amount': 0,
                        'net_amount': 0
                    }
                
                if row.txn_type == 'in':
                    summary[key]['inflow_count'] = row.txn_count
                    summary[key]['inflow_amount'] = row.total_amount
                else:
                    summary[key]['outflow_count'] = row.txn_count
                    summary[key]['outflow_amount'] = abs(row.total_amount)
                
                summary[key]['net_amount'] = (
                    summary[key]['inflow_amount'] - summary[key]['outflow_amount']
                )
            
            return list(summary.values())
